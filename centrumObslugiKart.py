from openpyxl import Workbook, load_workbook
from firma import FirmaTransportowa, Sklep, ZakladUslugowy
from bank import Bank

class CentrumObslugiKart:
    def __init__(self):
        self.__archiwum = []
        self.__lista_bankow = []
        self.__lista_firm = []
        self.wb = Workbook()

    def poczatek_pliku(self):

        ws = self.wb.active
        ws.append(["Nazwa banku","NIP firmy","numer karty","imie","nazwisko","kwota"])
        self.wb.save("Archiwum.xlsx")

    def dodajFirme(self, rodzaj, nazwa, NIP, nr_konta, saldo):
        if rodzaj == "sklep":
            self.__lista_firm.append(Sklep(nazwa, NIP, nr_konta, saldo))
        elif rodzaj == "zaklad uslugowy":
            self.__lista_firm.append(ZakladUslugowy(nazwa, NIP, nr_konta, saldo))
        elif rodzaj == "firma transportowa":
            self.__lista_firm.append(FirmaTransportowa(nazwa, NIP, nr_konta, saldo))

    def usunFirme(self, NIP):
        for firma in self.__lista_firm:
            firma_NIP = getattr(firma,'NIP',None)
            if firma_NIP == NIP:
                self.__lista_firm.remove(firma)
                break

    def przegladFirm(self):
        return self.__lista_firm


    def dodajBank(self, nazwa):
        bank = Bank(nazwa)
        self.__lista_bankow.append(bank)

    def usunBank(self,nazwa):
        for bank in self.__lista_bankow:
            bank_nazwa = getattr(bank,'nazwa')
            if bank_nazwa == nazwa:
                self.__lista_bankow.remove(bank)

    def przegladBankow(self):
        return self.__lista_bankow

    def platnosc(self, NIP, nr_karty, kwota, bank_klienta, bank_firmy):
        for i in range (0,len(self.__lista_bankow)):
            if self.__lista_bankow[i] == bank_klienta:
                znaleziony_bank = self.__lista_bankow[i]
                szukane_konto = znaleziony_bank.znajdzKonto(nr_karty)
                szukane_konto.saldo -= kwota
                break

        for i in range(0,len(self.__lista_bankow)):
            if self.__lista_bankow[i] == bank_firmy:
                znaleziony_bank_firmy = self.__lista_bankow[i]
                for firma in znaleziony_bank_firmy.getFirmy:
                    firma_NIP = getattr(firma,'NIP',None)
                    if firma_NIP == NIP:
                        szukana_firma = firma
                        szukane_konto_firmy = getattr(szukana_firma,'konto',None)
                        szukane_konto_firmy.saldo -= kwota
                        break

    def ZapiszDoPliku(self):
        wb = load_workbook('Archiwum.xlsx')
        arkusz = wb.active
        powtarzajace_sie_dane = []
        for wiersz in arkusz.iter_rows(values_only=True):
            for dane_excel in wiersz:
                if dane_excel:
                    for dane_archiwum in self.__archiwum:
                        if any(dane_excel in dane for dane in dane_archiwum.values()):
                            powtarzajace_sie_dane.append(dane_excel)

        brakujace_dane = []
        for dane_archiwum in self.__archiwum:
            if not any(dane in powtarzajace_sie_dane for dane in dane_archiwum.values()):
                brakujace_dane.append(dane_archiwum)

        for dane in brakujace_dane:
            arkusz.append(list(dane.values()))


    def OdczytZpliku(self):
        wb = load_workbook("Archiwum.xlsx")
        ws = wb.active
        for wiersz in ws.iter_rows(values_only=True):
            print(wiersz)

    def zarchiwizuj(self, bank, NIP, nr_karty, imie, nazwisko, kwota):
        platnosc = {
            "nazwa_banku": f"{getattr(bank, 'nazwa')}",
            "NIP_firmy": NIP,
            "nr_karty": nr_karty,
            "imie": imie,
            "nazwisko": nazwisko,
            "kwota": kwota,
        }
        self.__archiwum.append(platnosc)

    def przeszukiwanieArchiwum(self):
        while(True):
            print("----Archiwum----")
            print("1. historia platnosci firmy")
            print("2. historia platnosci przechodzacych przez dany bank")
            print("3. historia platnosci dana karta")
            print("4. historia platnosci danej osoby")
            print("5. historia platnosci wzgledem danej kwoty")
            print("6. Wyjdz z archiwum")
            wybor = input(print("Wybierz numer(1-5): "))
            match wybor:
                case 1:
                    NIP_firmy = input(print("Podaj NIP firmy: "))
                    for i in range(0,len(self.__archiwum)):
                        if self.__archiwum[i][1] == NIP_firmy:
                            print(self.__archiwum[i])

                case 2:
                    nazwa_banku = input(print("Podaj nazwe banku: "))
                    for i in range(0,len(self.__archiwum)):
                        if self.__archiwum[i][0] == nazwa_banku:
                            print(self.__archiwum[i])

                case 3:
                    nr_karty = input(print("Podaj numer karty: "))
                    for i in range(0, len(self.__archiwum)):
                        if self.__archiwum[i][2] == nr_karty:
                            print(self.__archiwum[i])

                case 4:
                    imie = input(print("Podaj imie osoby: "))
                    nazwisko = input(print("Podaj nazwisko osoby: "))
                    for i in range(0, len(self.__archiwum)):
                        if self.__archiwum[i][3] == imie and self.__archiwum[i][4] == nazwisko:
                            print(self.__archiwum[i])

                case 5:
                    kwota = input(print("Podaj kwote: "))
                    print("Wybierz co chcesz zrobic: ")
                    print("1. Znajdz historie platnosci rowna danej kwocie")
                    print("2. Znajdz historie platnosci od danej kwoty w gore")
                    print("3. Znajdz historie platnosci ponizej danej kwoty")
                    match wybor:
                        case 1:
                            for i in range(0,len(self.__archiwum)):
                                if self.__archiwum[i][5] == kwota:
                                    print(self.__archiwum[i])

                        case 2:
                            for i in range(0, len(self.__archiwum)):
                                if self.__archiwum[i][5] > kwota:
                                    print(self.__archiwum[i])

                        case 3:
                            for i in range(0, len(self.__archiwum)):
                                if self.__archiwum[i][5] < kwota:
                                    print(self.__archiwum[i])

                case 6:
                    break